"""
Utilidades Avanzadas de Excel para el Ecosistema de Emprendimiento

Este módulo proporciona una clase para generar y leer archivos Excel (.xlsx)
con funcionalidades avanzadas como múltiples hojas, estilos personalizados,
formateo de celdas, y más. Utiliza la librería openpyxl.

Author: Sistema de Emprendimiento
Version: 1.0.0
"""

import logging
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.chart import BarChart, LineChart, Reference, Series
from flask import Response, current_app

logger = logging.getLogger(__name__)

# Estilos predefinidos
HEADER_STYLE = NamedStyle(name="header_style")
HEADER_STYLE.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
HEADER_STYLE.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_STYLE.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Side(border_style="thin", color="000000")
HEADER_STYLE.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

DATA_STYLE = NamedStyle(name="data_style")
DATA_STYLE.font = Font(name="Calibri", size=11)
DATA_STYLE.alignment = Alignment(vertical="top", wrap_text=False)

CURRENCY_STYLE_COP = NamedStyle(name="currency_cop", number_format='"$"#,##0')
CURRENCY_STYLE_USD = NamedStyle(name="currency_usd", number_format='"$"#,##0.00')
DATE_STYLE = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
DATETIME_STYLE = NamedStyle(name="datetime_style", number_format="YYYY-MM-DD HH:MM:SS")
PERCENTAGE_STYLE = NamedStyle(name="percentage_style", number_format="0.00%")


class ExcelBuilder:
    """
    Clase para construir archivos Excel (.xlsx) con funcionalidades avanzadas.
    """

    def __init__(self):
        self.workbook = Workbook()
        # Eliminar la hoja por defecto si se va a trabajar con hojas nombradas
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # Registrar estilos nombrados
        if "header_style" not in self.workbook.named_styles:
            self.workbook.add_named_style(HEADER_STYLE)
        if "data_style" not in self.workbook.named_styles:
            self.workbook.add_named_style(DATA_STYLE)
        if "currency_cop" not in self.workbook.named_styles:
            self.workbook.add_named_style(CURRENCY_STYLE_COP)
        if "currency_usd" not in self.workbook.named_styles:
            self.workbook.add_named_style(CURRENCY_STYLE_USD)
        if "date_style" not in self.workbook.named_styles:
            self.workbook.add_named_style(DATE_STYLE)
        if "datetime_style" not in self.workbook.named_styles:
            self.workbook.add_named_style(DATETIME_STYLE)
        if "percentage_style" not in self.workbook.named_styles:
            self.workbook.add_named_style(PERCENTAGE_STYLE)

    def add_sheet_from_list_of_dicts(
        self,
        sheet_name: str,
        data: List[Dict[str, Any]],
        headers: Optional[List[str]] = None,
        column_styles: Optional[Dict[str, str]] = None,
        auto_adjust_column_widths: bool = True
    ) -> Worksheet:
        """
        Agrega una nueva hoja al workbook a partir de una lista de diccionarios.

        Args:
            sheet_name: Nombre de la hoja.
            data: Lista de diccionarios.
            headers: Lista opcional de headers. Si es None, se infieren de las claves del primer dict.
            column_styles: Diccionario opcional para mapear headers a estilos nombrados.
            auto_adjust_column_widths: Si ajustar automáticamente el ancho de las columnas.

        Returns:
            Worksheet: La hoja de cálculo creada.
        """
        if not data:
            ws = self.workbook.create_sheet(title=sheet_name)
            if headers:
                ws.append(headers)
                for cell in ws[1]:
                    cell.style = "header_style"
            return ws

        ws = self.workbook.create_sheet(title=sheet_name)
        
        if headers is None:
            headers = list(data[0].keys())
        
        # Escribir headers
        ws.append(headers)
        for cell in ws[1]:
            cell.style = "header_style"

        # Escribir datos
        for row_data in data:
            row_values = [row_data.get(header, "") for header in headers]
            ws.append(row_values)

        # Aplicar estilos a columnas de datos
        if column_styles:
            for col_idx, header in enumerate(headers, 1):
                if header in column_styles:
                    style_name = column_styles[header]
                    if style_name in self.workbook.named_styles:
                        for row_idx in range(2, ws.max_row + 1):
                            ws.cell(row=row_idx, column=col_idx).style = style_name
                    else:
                        logger.warning(f"Estilo '{style_name}' no encontrado para columna '{header}'.")
        else: # Aplicar estilo de datos por defecto
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.style = "data_style"


        if auto_adjust_column_widths:
            self.auto_adjust_column_widths(ws, headers, data)
            
        return ws

    def add_sheet_from_dataframe(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        index: bool = False,
        header: bool = True,
        column_styles: Optional[Dict[str, str]] = None,
        auto_adjust_column_widths: bool = True
    ) -> Worksheet:
        """
        Agrega una nueva hoja al workbook a partir de un DataFrame de pandas.

        Args:
            sheet_name: Nombre de la hoja.
            df: DataFrame de pandas.
            index: Si incluir el índice del DataFrame.
            header: Si incluir los headers del DataFrame.
            column_styles: Diccionario opcional para mapear headers a estilos nombrados.
            auto_adjust_column_widths: Si ajustar automáticamente el ancho de las columnas.

        Returns:
            Worksheet: La hoja de cálculo creada.
        """
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Convertir DataFrame a filas, incluyendo headers si es necesario
        for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=header), 1):
            ws.append(row)
            if r_idx == 1 and header: # Aplicar estilo a la fila de encabezado
                for cell in ws[1]:
                    cell.style = "header_style"
        
        # Aplicar estilos a columnas de datos
        if column_styles and header:
            headers_list = list(df.columns)
            if index:
                headers_list = [df.index.name or "index"] + headers_list

            for col_idx, header_name in enumerate(headers_list, 1):
                if header_name in column_styles:
                    style_name = column_styles[header_name]
                    if style_name in self.workbook.named_styles:
                        for row_idx in range(2, ws.max_row + 1): # Empezar desde la fila 2 si hay header
                            ws.cell(row=row_idx, column=col_idx).style = style_name
                    else:
                        logger.warning(f"Estilo '{style_name}' no encontrado para columna '{header_name}'.")
        elif header: # Aplicar estilo de datos por defecto
             for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).style = "data_style"


        if auto_adjust_column_widths:
            self.auto_adjust_column_widths_from_dataframe(ws, df, index=index, header=header)
            
        return ws

    def auto_adjust_column_widths(self, ws: Worksheet, headers: List[str], data: List[Dict[str, Any]]):
        """Ajusta automáticamente el ancho de las columnas basado en el contenido."""
        column_widths = {}
        for col_idx, header in enumerate(headers, 1):
            column_widths[col_idx] = len(str(header))
            for row_data in data:
                cell_value = row_data.get(header, "")
                column_widths[col_idx] = max(column_widths[col_idx], len(str(cell_value)))
        
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 2 # Pequeño padding

    def auto_adjust_column_widths_from_dataframe(self, ws: Worksheet, df: pd.DataFrame, index: bool, header: bool):
        """Ajusta el ancho de las columnas basado en un DataFrame."""
        column_names = list(df.columns)
        if index:
            column_names.insert(0, df.index.name or "index")

        for i, column_name in enumerate(column_names, 1):
            max_length = 0
            if header:
                max_length = len(str(column_name))
            
            # Iterar sobre los datos de la columna
            column_data = df.iloc[:, i-1] if not index or i > 1 else df.index
            for cell_value in column_data:
                try:
                    cell_len = len(str(cell_value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass # Ignorar errores de conversión
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width


    def add_chart(
        self,
        ws: Worksheet,
        chart_type: str, # 'bar', 'line'
        title: str,
        data_start_cell: str, # ej. "B2"
        data_end_cell: str,   # ej. "B10"
        categories_start_cell: str, # ej. "A2"
        categories_end_cell: str,   # ej. "A10"
        chart_position: str = "E2", # ej. "E2"
        x_axis_title: Optional[str] = None,
        y_axis_title: Optional[str] = None
    ):
        """Agrega un gráfico a una hoja."""
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        else:
            logger.warning(f"Tipo de gráfico '{chart_type}' no soportado.")
            return

        chart.title = title
        if x_axis_title:
            chart.x_axis.title = x_axis_title
        if y_axis_title:
            chart.y_axis.title = y_axis_title

        data = Reference(ws, range_string=f"{ws.title}!{data_start_cell}:{data_end_cell}")
        cats = Reference(ws, range_string=f"{ws.title}!{categories_start_cell}:{categories_end_cell}")
        
        chart.add_data(data, titles_from_data=True) # Asume que la primera celda del rango de datos es el título de la serie
        chart.set_categories(cats)
        
        ws.add_chart(chart, chart_position)

    def get_bytes(self) -> BytesIO:
        """
        Guarda el workbook en un objeto BytesIO.
        
        Returns:
            BytesIO: El workbook como bytes.
        """
        excel_bytes = BytesIO()
        self.workbook.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes

    def save_to_file(self, filename: str):
        """
        Guarda el workbook en un archivo.
        
        Args:
            filename: Nombre del archivo (ej. "reporte.xlsx").
        """
        self.workbook.save(filename)
        logger.info(f"Workbook guardado en {filename}")


class ExcelReader:
    """
    Clase para leer y parsear archivos Excel (.xlsx).
    """

    def __init__(self, file_path_or_bytes: Union[str, BytesIO]):
        """
        Inicializa el lector de Excel.

        Args:
            file_path_or_bytes: Ruta al archivo Excel o un objeto BytesIO.
        """
        try:
            self.workbook = openpyxl.load_workbook(file_path_or_bytes, data_only=True) # data_only=True para obtener valores de fórmulas
        except Exception as e:
            logger.error(f"Error cargando workbook Excel: {e}")
            raise ValueError(f"No se pudo cargar el archivo Excel: {e}")

    def get_sheet_names(self) -> List[str]:
        """Retorna los nombres de todas las hojas en el workbook."""
        return self.workbook.sheetnames

    def read_sheet_to_list_of_dicts(
        self,
        sheet_name: Optional[str] = None,
        header_row: int = 1
    ) -> List[Dict[str, Any]]:
        """
        Lee una hoja específica y la convierte en una lista de diccionarios.

        Args:
            sheet_name: Nombre de la hoja. Si es None, usa la primera hoja activa.
            header_row: Número de la fila que contiene los encabezados (1-indexed).

        Returns:
            Lista de diccionarios, donde cada diccionario representa una fila.
        """
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Hoja '{sheet_name}' no encontrada en el archivo.")
            ws = self.workbook[sheet_name]
        else:
            ws = self.workbook.active

        data = []
        headers = [cell.value for cell in ws[header_row]]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            row_data = {}
            empty_row = True
            for col_idx, cell in enumerate(row):
                value = cell.value
                # Convertir datetime a ISO string si es necesario
                if isinstance(value, datetime):
                    value = value.isoformat()
                row_data[headers[col_idx]] = value
                if value is not None:
                    empty_row = False
            if not empty_row:
                data.append(row_data)
                
        return data

    def read_sheet_to_dataframe(
        self,
        sheet_name: Optional[str] = None,
        header_row: int = 0 # pandas usa 0-indexed para header
    ) -> pd.DataFrame:
        """
        Lee una hoja específica y la convierte en un DataFrame de pandas.

        Args:
            sheet_name: Nombre de la hoja. Si es None, usa la primera hoja activa.
            header_row: Número de la fila que contiene los encabezados (0-indexed para pandas).

        Returns:
            DataFrame de pandas con los datos de la hoja.
        """
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Hoja '{sheet_name}' no encontrada en el archivo.")
            ws = self.workbook[sheet_name]
        else:
            ws = self.workbook.active
            sheet_name = ws.title # Necesario para pandas.read_excel si pasamos BytesIO

        # Convertir la hoja de openpyxl a un formato que pandas pueda leer directamente (BytesIO)
        # Esto es más robusto que iterar filas manualmente para DataFrames.
        excel_bytes = BytesIO()
        temp_wb = Workbook()
        if "Sheet" in temp_wb.sheetnames: # Eliminar hoja por defecto si existe
            temp_wb.remove(temp_wb["Sheet"])
        temp_ws = temp_wb.create_sheet(title=sheet_name)

        for row in ws.iter_rows():
            temp_ws.append([cell.value for cell in row])
        
        temp_wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        df = pd.read_excel(excel_bytes, sheet_name=sheet_name, header=header_row)
        return df


def generate_excel_response(
    data_sheets: Dict[str, List[Dict[str, Any]]],
    filename_base: str,
    sheet_styles: Optional[Dict[str, Dict[str, str]]] = None
) -> Response:
    """
    Genera una respuesta Flask con un archivo Excel.

    Args:
        data_sheets: Diccionario donde las claves son nombres de hojas y los valores son listas de diccionarios.
        filename_base: Nombre base para el archivo (sin extensión).
        sheet_styles: Opcional. Diccionario donde las claves son nombres de hojas y los valores son
                      diccionarios mapeando headers de columna a estilos nombrados.

    Returns:
        Response: Objeto Response de Flask con el archivo Excel.
    """
    builder = ExcelBuilder()
    
    for sheet_name, data in data_sheets.items():
        column_styles_for_sheet = sheet_styles.get(sheet_name) if sheet_styles else None
        builder.add_sheet_from_list_of_dicts(sheet_name, data, column_styles=column_styles_for_sheet)

    excel_bytes = builder.get_bytes()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{filename_base}_{timestamp}.xlsx"
    
    return Response(
        excel_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

# Ejemplo de uso (se puede remover o comentar en producción)
if __name__ == '__main__':
    # Ejemplo de creación
    builder = ExcelBuilder()
    
    data1 = [
        {'Nombre': 'Juan Pérez', 'Edad': 30, 'Ciudad': 'Bogotá', 'Ingresos': 5000000, 'Registro': datetime(2023,1,15,10,30)},
        {'Nombre': 'Ana García', 'Edad': 25, 'Ciudad': 'Medellín', 'Ingresos': 3500000.50, 'Registro': datetime(2023,2,20,14,45)},
    ]
    styles1 = {
        'Ingresos': 'currency_cop',
        'Registro': 'datetime_style'
    }
    builder.add_sheet_from_list_of_dicts("Usuarios", data1, column_styles=styles1)

    df_data = {
        'Producto': ['A', 'B', 'C'],
        'Ventas_USD': [1200.75, 2500.00, 850.50],
        'Fecha_Reporte': [date(2023,3,1), date(2023,3,5), date(2023,3,10)],
        'Rentabilidad': [0.15, 0.22, 0.18]
    }
    df = pd.DataFrame(df_data)
    styles2 = {
        'Ventas_USD': 'currency_usd',
        'Fecha_Reporte': 'date_style',
        'Rentabilidad': 'percentage_style'
    }
    ws_productos = builder.add_sheet_from_dataframe("Productos", df, column_styles=styles2)
    
    # Agregar un gráfico
    builder.add_chart(
        ws_productos, 
        chart_type='bar', 
        title="Ventas por Producto",
        data_start_cell="B2", data_end_cell="B4", # Datos de Ventas_USD
        categories_start_cell="A2", categories_end_cell="A4", # Datos de Producto
        chart_position="E2",
        y_axis_title="Ventas (USD)"
    )

    builder.save_to_file("reporte_avanzado.xlsx")
    print("Archivo 'reporte_avanzado.xlsx' generado.")

    # Ejemplo de lectura
    try:
        reader = ExcelReader("reporte_avanzado.xlsx")
        print("\nNombres de las hojas:", reader.get_sheet_names())
        
        usuarios_data = reader.read_sheet_to_list_of_dicts("Usuarios")
        print("\nDatos de Usuarios (lista de dicts):")
        for row in usuarios_data:
            print(row)
            
        productos_df = reader.read_sheet_to_dataframe("Productos")
        print("\nDatos de Productos (DataFrame):")
        print(productos_df)
        
    except Exception as e:
        print(f"Error leyendo archivo: {e}")