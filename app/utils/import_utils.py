"""
Utilidades de Importación - Ecosistema de Emprendimiento
=======================================================

Este módulo proporciona un conjunto completo de utilidades para la importación
masiva de datos desde diferentes fuentes (CSV, Excel, JSON, XML) específicamente
optimizado para el ecosistema de emprendimiento colombiano.

Características principales:
- Soporte múltiples formatos: CSV, Excel, JSON, XML, TSV
- Validación de datos en tiempo real
- Mapeo flexible de campos
- Procesamiento por lotes con progress tracking
- Preview de datos antes de importación
- Transformaciones automáticas de datos
- Rollback automático en errores críticos
- Logging detallado de operaciones
- Utilidades específicas para emprendimiento
- Manejo de archivos grandes (streaming)
- Detección automática de encoding
- Limpieza y normalización de datos

Uso básico:
-----------
    from app.utils.import_utils import ImportManager, import_from_csv
    
    # Importación simple desde CSV
    result = import_from_csv('emprendedores.csv', 'Entrepreneur')
    
    # Importación avanzada con validaciones
    manager = ImportManager()
    result = manager.import_data(
        'proyectos.xlsx', 
        target_model='Project',
        field_mapping={'nombre': 'name', 'sector': 'business_sector'},
        validate_data=True
    )

Formatos soportados:
-------------------
- CSV (Comma Separated Values)
- Excel (.xlsx, .xls)
- JSON (JavaScript Object Notation)
- XML (eXtensible Markup Language)
- TSV (Tab Separated Values)
- JSONL (JSON Lines)
"""

import csv
import json
import logging
import mimetypes
import os
import tempfile
import time
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import (
    Any, Dict, List, Optional, Union, Tuple, Callable, 
    Iterator, Generator, Type, IO, Set
)
from dataclasses import dataclass, field
from enum import Enum
import re
import chardet
import pandas as pd
from io import StringIO, BytesIO

# Imports opcionales con manejo de errores
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import xml.etree.ElementTree as ET
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False

try:
    import sqlalchemy
    from sqlalchemy.orm import Session
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    SQLALCHEMY_AVAILABLE = False

# Configurar logger
logger = logging.getLogger(__name__)

# ==============================================================================
# CONFIGURACIÓN Y CONSTANTES
# ==============================================================================

# Configuración de importación
IMPORT_CONFIG = {
    'max_file_size': 50 * 1024 * 1024,  # 50MB
    'batch_size': 1000,
    'max_rows': 100000,
    'max_preview_rows': 100,
    'default_encoding': 'utf-8',
    'encoding_detection_sample_size': 10000,
    'supported_encodings': ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1'],
    'csv_delimiter_detection': [',', ';', '\t', '|'],
    'date_formats': [
        '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
        '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S.%f'
    ],
    'boolean_values': {
        'true': [True, 'true', 'True', 'TRUE', '1', 'yes', 'Yes', 'YES', 'sí', 'Sí', 'SÍ'],
        'false': [False, 'false', 'False', 'FALSE', '0', 'no', 'No', 'NO']
    },
    'null_values': ['', 'null', 'NULL', 'None', 'NONE', 'N/A', 'n/a', 'NA', '#N/A'],
    'chunk_size': 1000,  # Para archivos grandes
    'temp_dir': tempfile.gettempdir(),
    'validate_on_import': True,
    'rollback_on_error': True,
    'progress_callback_interval': 100,
}

# Tipos de archivo soportados
class FileType(Enum):
    """Tipos de archivo soportados."""
    CSV = 'csv'
    EXCEL = 'excel'
    JSON = 'json'
    XML = 'xml'
    TSV = 'tsv'
    JSONL = 'jsonl'

class ImportStatus(Enum):
    """Estados de importación."""
    PENDING = 'pending'
    PROCESSING = 'processing'
    COMPLETED = 'completed'
    FAILED = 'failed'
    CANCELLED = 'cancelled'

class ValidationLevel(Enum):
    """Niveles de validación."""
    NONE = 'none'
    BASIC = 'basic'
    STRICT = 'strict'
    CUSTOM = 'custom'

# Mapeos específicos para emprendimiento
ENTREPRENEUR_FIELD_MAPPING = {
    'nombre': 'name',
    'apellido': 'last_name',
    'email': 'email',
    'telefono': 'phone',
    'celular': 'mobile_phone',
    'empresa': 'company_name',
    'sector': 'business_sector',
    'ciudad': 'city',
    'pais': 'country',
    'experiencia': 'experience_years',
    'educacion': 'education_level',
    'linkedin': 'linkedin_url',
    'website': 'website_url',
    'descripcion': 'description',
    'estado': 'status',
    'fecha_registro': 'created_at',
    'activo': 'is_active'
}

PROJECT_FIELD_MAPPING = {
    'nombre': 'name',
    'descripcion': 'description',
    'sector': 'business_sector',
    'etapa': 'stage',
    'presupuesto': 'budget',
    'moneda': 'currency',
    'fecha_inicio': 'start_date',
    'fecha_fin': 'end_date',
    'emprendedor': 'entrepreneur_id',
    'mentor': 'mentor_id',
    'estado': 'status',
    'progreso': 'progress_percentage',
    'objetivo': 'objective',
    'mercado_objetivo': 'target_market',
    'propuesta_valor': 'value_proposition',
    'equipo': 'team_size',
    'inversion_requerida': 'required_investment',
    'inversion_actual': 'current_investment'
}

BUSINESS_SECTORS = {
    'tecnologia', 'agricultura', 'manufactura', 'servicios', 'comercio',
    'turismo', 'educacion', 'salud', 'construccion', 'transporte',
    'energia', 'finanzas', 'alimentario', 'textil', 'creativos',
    'mineria', 'telecomunicaciones', 'consultoria', 'otro'
}

# ==============================================================================
# CLASES DE DATOS Y EXCEPCIONES
# ==============================================================================

class ImportError(Exception):
    """Excepción base para errores de importación."""
    pass

class FileValidationError(ImportError):
    """Error de validación de archivo."""
    pass

class DataValidationError(ImportError):
    """Error de validación de datos."""
    pass

class MappingError(ImportError):
    """Error en mapeo de campos."""
    pass

class ProcessingError(ImportError):
    """Error en procesamiento de datos."""
    pass

@dataclass
class ImportResult:
    """Resultado de una operación de importación."""
    status: ImportStatus
    total_rows: int = 0
    processed_rows: int = 0
    successful_rows: int = 0
    failed_rows: int = 0
    errors: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[Dict[str, Any]] = field(default_factory=list)
    imported_data: List[Dict[str, Any]] = field(default_factory=list)
    processing_time: Optional[float] = None
    file_info: Optional[Dict[str, Any]] = None
    field_mapping: Optional[Dict[str, str]] = None
    validation_summary: Optional[Dict[str, Any]] = None
    
    def add_error(self, row_number: int, error_message: str, field: str = None):
        """Añade un error al resultado."""
        self.errors.append({
            'row': row_number,
            'message': error_message,
            'field': field,
            'timestamp': datetime.now().isoformat()
        })
        self.failed_rows += 1
    
    def add_warning(self, row_number: int, warning_message: str, field: str = None):
        """Añade una advertencia al resultado."""
        self.warnings.append({
            'row': row_number,
            'message': warning_message,
            'field': field,
            'timestamp': datetime.now().isoformat()
        })
    
    def get_success_rate(self) -> float:
        """Calcula la tasa de éxito."""
        if self.total_rows == 0:
            return 0.0
        return (self.successful_rows / self.total_rows) * 100

@dataclass
class FieldMapping:
    """Configuración de mapeo de campos."""
    source_field: str
    target_field: str
    data_type: str = 'string'
    required: bool = False
    default_value: Any = None
    validator: Optional[Callable] = None
    transformer: Optional[Callable] = None
    description: str = ''

@dataclass
class ImportConfig:
    """Configuración de importación."""
    file_path: str
    target_model: Optional[str] = None
    field_mappings: List[FieldMapping] = field(default_factory=list)
    validation_level: ValidationLevel = ValidationLevel.BASIC
    batch_size: int = 1000
    max_rows: Optional[int] = None
    encoding: Optional[str] = None
    delimiter: Optional[str] = None
    header_row: int = 0
    skip_rows: int = 0
    preview_only: bool = False
    rollback_on_error: bool = True
    progress_callback: Optional[Callable] = None
    custom_validators: List[Callable] = field(default_factory=list)
    transformation_rules: Dict[str, Callable] = field(default_factory=dict)

# ==============================================================================
# UTILIDADES DE DETECCIÓN Y VALIDACIÓN DE ARCHIVOS
# ==============================================================================

def detect_file_type(file_path: str) -> FileType:
    """
    Detecta el tipo de archivo basado en la extensión y contenido.
    
    Args:
        file_path: Ruta del archivo
        
    Returns:
        Tipo de archivo detectado
        
    Examples:
        >>> file_type = detect_file_type('data.csv')
        >>> print(file_type)  # FileType.CSV
    """
    if not os.path.exists(file_path):
        raise FileValidationError(f"Archivo no encontrado: {file_path}")
    
    # Detectar por extensión
    extension = Path(file_path).suffix.lower()
    
    extension_mapping = {
        '.csv': FileType.CSV,
        '.tsv': FileType.TSV,
        '.txt': FileType.CSV,  # Asumir CSV para .txt
        '.xlsx': FileType.EXCEL,
        '.xls': FileType.EXCEL,
        '.json': FileType.JSON,
        '.jsonl': FileType.JSONL,
        '.xml': FileType.XML
    }
    
    if extension in extension_mapping:
        return extension_mapping[extension]
    
    # Detectar por contenido MIME
    mime_type, _ = mimetypes.guess_type(file_path)
    
    if mime_type:
        if 'csv' in mime_type or 'text/plain' in mime_type:
            return FileType.CSV
        elif 'excel' in mime_type or 'spreadsheet' in mime_type:
            return FileType.EXCEL
        elif 'json' in mime_type:
            return FileType.JSON
        elif 'xml' in mime_type:
            return FileType.XML
    
    # Fallback: intentar detectar por contenido
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            sample = f.read(1024)
            
        if sample.strip().startswith('{') or sample.strip().startswith('['):
            return FileType.JSON
        elif sample.strip().startswith('<'):
            return FileType.XML
        else:
            return FileType.CSV
            
    except Exception:
        raise FileValidationError(f"No se pudo detectar el tipo de archivo: {file_path}")

def detect_encoding(file_path: str, sample_size: int = None) -> str:
    """
    Detecta la codificación de un archivo.
    
    Args:
        file_path: Ruta del archivo
        sample_size: Tamaño de muestra para detección
        
    Returns:
        Codificación detectada
    """
    if sample_size is None:
        sample_size = IMPORT_CONFIG['encoding_detection_sample_size']
    
    try:
        with open(file_path, 'rb') as f:
            raw_data = f.read(sample_size)
        
        result = chardet.detect(raw_data)
        encoding = result['encoding']
        confidence = result['confidence']
        
        logger.debug(f"Encoding detectado: {encoding} (confianza: {confidence:.2f})")
        
        # Validar que el encoding sea soportado
        if encoding and encoding.lower() in IMPORT_CONFIG['supported_encodings']:
            return encoding
        
        # Fallback a UTF-8
        return IMPORT_CONFIG['default_encoding']
        
    except Exception as e:
        logger.warning(f"Error detectando encoding: {e}")
        return IMPORT_CONFIG['default_encoding']

def detect_csv_delimiter(file_path: str, encoding: str = 'utf-8') -> str:
    """
    Detecta el delimitador de un archivo CSV.
    
    Args:
        file_path: Ruta del archivo
        encoding: Codificación del archivo
        
    Returns:
        Delimitador detectado
    """
    try:
        with open(file_path, 'r', encoding=encoding) as f:
            sample = f.read(4096)
        
        # Usar csv.Sniffer para detectar
        sniffer = csv.Sniffer()
        delimiter = sniffer.sniff(sample, delimiters=',;\t|').delimiter
        
        logger.debug(f"Delimitador detectado: '{delimiter}'")
        return delimiter
        
    except Exception:
        # Fallback: contar ocurrencias de delimitadores comunes
        delimiter_counts = {}
        
        for delimiter in IMPORT_CONFIG['csv_delimiter_detection']:
            count = sample.count(delimiter)
            delimiter_counts[delimiter] = count
        
        # Retornar el más común
        best_delimiter = max(delimiter_counts, key=delimiter_counts.get)
        
        if delimiter_counts[best_delimiter] > 0:
            return best_delimiter
        
        # Fallback final
        return ','

def validate_file(file_path: str) -> Dict[str, Any]:
    """
    Valida un archivo antes de la importación.
    
    Args:
        file_path: Ruta del archivo
        
    Returns:
        Información de validación del archivo
        
    Raises:
        FileValidationError: Si el archivo no es válido
    """
    if not os.path.exists(file_path):
        raise FileValidationError(f"Archivo no encontrado: {file_path}")
    
    file_info = {
        'path': file_path,
        'size': os.path.getsize(file_path),
        'name': os.path.basename(file_path),
        'extension': Path(file_path).suffix.lower()
    }
    
    # Validar tamaño
    max_size = IMPORT_CONFIG['max_file_size']
    if file_info['size'] > max_size:
        raise FileValidationError(
            f"Archivo demasiado grande: {file_info['size']} bytes "
            f"(máximo: {max_size} bytes)"
        )
    
    # Detectar tipo y propiedades
    try:
        file_info['type'] = detect_file_type(file_path)
        file_info['encoding'] = detect_encoding(file_path)
        
        if file_info['type'] in [FileType.CSV, FileType.TSV]:
            file_info['delimiter'] = detect_csv_delimiter(file_path, file_info['encoding'])
        
        # Validar que el archivo se puede leer
        _validate_file_readability(file_path, file_info)
        
        logger.info(f"Archivo validado: {file_info}")
        return file_info
        
    except Exception as e:
        raise FileValidationError(f"Error validando archivo: {e}")

def _validate_file_readability(file_path: str, file_info: Dict[str, Any]):
    """Valida que el archivo se puede leer correctamente."""
    try:
        file_type = file_info['type']
        encoding = file_info['encoding']
        
        if file_type == FileType.CSV:
            delimiter = file_info.get('delimiter', ',')
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f, delimiter=delimiter)
                next(reader)  # Intentar leer primera línea
                
        elif file_type == FileType.EXCEL:
            if not OPENPYXL_AVAILABLE:
                raise FileValidationError("openpyxl no está disponible para archivos Excel")
            load_workbook(file_path, read_only=True)
            
        elif file_type == FileType.JSON:
            with open(file_path, 'r', encoding=encoding) as f:
                json.load(f)
                
        elif file_type == FileType.XML:
            if not XML_AVAILABLE:
                raise FileValidationError("xml.etree.ElementTree no está disponible")
            ET.parse(file_path)
            
    except Exception as e:
        raise FileValidationError(f"Archivo no se puede leer: {e}")

# ==============================================================================
# LECTORES DE ARCHIVOS
# ==============================================================================

class BaseFileReader:
    """Clase base para lectores de archivos."""
    
    def __init__(self, file_path: str, encoding: str = 'utf-8'):
        self.file_path = file_path
        self.encoding = encoding
        self.file_info = validate_file(file_path)
    
    def read(self, max_rows: Optional[int] = None) -> Iterator[Dict[str, Any]]:
        """Lee el archivo y retorna un iterador de diccionarios."""
        raise NotImplementedError
    
    def get_headers(self) -> List[str]:
        """Obtiene las cabeceras del archivo."""
        raise NotImplementedError
    
    def count_rows(self) -> int:
        """Cuenta el número de filas del archivo."""
        raise NotImplementedError

class CSVReader(BaseFileReader):
    """Lector para archivos CSV."""
    
    def __init__(self, file_path: str, encoding: str = None, delimiter: str = None):
        super().__init__(file_path, encoding or detect_encoding(file_path))
        self.delimiter = delimiter or detect_csv_delimiter(file_path, self.encoding)
    
    def read(self, max_rows: Optional[int] = None) -> Iterator[Dict[str, Any]]:
        """Lee archivo CSV línea por línea."""
        try:
            with open(self.file_path, 'r', encoding=self.encoding) as f:
                reader = csv.DictReader(f, delimiter=self.delimiter)
                
                for row_num, row in enumerate(reader, 1):
                    if max_rows and row_num > max_rows:
                        break
                    
                    # Limpiar valores None y espacios
                    cleaned_row = {}
                    for key, value in row.items():
                        if value is not None:
                            cleaned_value = str(value).strip()
                            cleaned_row[key] = cleaned_value if cleaned_value else None
                        else:
                            cleaned_row[key] = None
                    
                    yield cleaned_row
                    
        except Exception as e:
            raise ProcessingError(f"Error leyendo CSV: {e}")
    
    def get_headers(self) -> List[str]:
        """Obtiene cabeceras del CSV."""
        try:
            with open(self.file_path, 'r', encoding=self.encoding) as f:
                reader = csv.reader(f, delimiter=self.delimiter)
                headers = next(reader)
                return [header.strip() for header in headers]
        except Exception as e:
            raise ProcessingError(f"Error obteniendo cabeceras CSV: {e}")
    
    def count_rows(self) -> int:
        """Cuenta filas del CSV (excluyendo cabecera)."""
        try:
            with open(self.file_path, 'r', encoding=self.encoding) as f:
                reader = csv.reader(f, delimiter=self.delimiter)
                next(reader)  # Saltar cabecera
                return sum(1 for _ in reader)
        except Exception as e:
            raise ProcessingError(f"Error contando filas CSV: {e}")

class ExcelReader(BaseFileReader):
    """Lector para archivos Excel."""
    
    def __init__(self, file_path: str, sheet_name: str = None):
        super().__init__(file_path)
        self.sheet_name = sheet_name
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está disponible para archivos Excel")
    
    def read(self, max_rows: Optional[int] = None) -> Iterator[Dict[str, Any]]:
        """Lee archivo Excel línea por línea."""
        try:
            workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            
            # Seleccionar hoja
            if self.sheet_name:
                if self.sheet_name in workbook.sheetnames:
                    worksheet = workbook[self.sheet_name]
                else:
                    raise ProcessingError(f"Hoja '{self.sheet_name}' no encontrada")
            else:
                worksheet = workbook.active
            
            # Obtener cabeceras de la primera fila
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value if cell.value else f"Column_{len(headers)+1}")
            
            # Leer datos
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                if max_rows and row_num > max_rows:
                    break
                
                # Crear diccionario con cabeceras
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        # Convertir valores especiales de Excel
                        if value is not None:
                            if isinstance(value, (int, float, Decimal)):
                                row_dict[headers[i]] = value
                            elif isinstance(value, datetime):
                                row_dict[headers[i]] = value.isoformat()
                            else:
                                row_dict[headers[i]] = str(value).strip()
                        else:
                            row_dict[headers[i]] = None
                
                yield row_dict
                
        except Exception as e:
            raise ProcessingError(f"Error leyendo Excel: {e}")
    
    def get_headers(self) -> List[str]:
        """Obtiene cabeceras del Excel."""
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            worksheet = workbook.active if not self.sheet_name else workbook[self.sheet_name]
            
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value if cell.value else f"Column_{len(headers)+1}")
            
            return headers
            
        except Exception as e:
            raise ProcessingError(f"Error obteniendo cabeceras Excel: {e}")
    
    def count_rows(self) -> int:
        """Cuenta filas del Excel (excluyendo cabecera)."""
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            worksheet = workbook.active if not self.sheet_name else workbook[self.sheet_name]
            return worksheet.max_row - 1  # Restar cabecera
        except Exception as e:
            raise ProcessingError(f"Error contando filas Excel: {e}")

class JSONReader(BaseFileReader):
    """Lector para archivos JSON."""
    
    def read(self, max_rows: Optional[int] = None) -> Iterator[Dict[str, Any]]:
        """Lee archivo JSON."""
        try:
            with open(self.file_path, 'r', encoding=self.encoding) as f:
                data = json.load(f)
            
            # Manejar diferentes estructuras JSON
            if isinstance(data, list):
                items = data
            elif isinstance(data, dict):
                if 'data' in data:
                    items = data['data']
                elif 'items' in data:
                    items = data['items']
                else:
                    items = [data]
            else:
                items = [data]
            
            for i, item in enumerate(items):
                if max_rows and i >= max_rows:
                    break
                
                if isinstance(item, dict):
                    yield item
                else:
                    yield {'value': item}
                    
        except Exception as e:
            raise ProcessingError(f"Error leyendo JSON: {e}")
    
    def get_headers(self) -> List[str]:
        """Obtiene cabeceras del JSON."""
        try:
            # Leer una muestra para obtener claves
            sample_data = list(self.read(max_rows=10))
            
            if not sample_data:
                return []
            
            # Obtener todas las claves únicas
            headers = set()
            for item in sample_data:
                if isinstance(item, dict):
                    headers.update(item.keys())
            
            return sorted(list(headers))
            
        except Exception as e:
            raise ProcessingError(f"Error obteniendo cabeceras JSON: {e}")
    
    def count_rows(self) -> int:
        """Cuenta elementos del JSON."""
        return sum(1 for _ in self.read())

class JSONLReader(BaseFileReader):
    """Lector para archivos JSONL (JSON Lines)."""
    
    def read(self, max_rows: Optional[int] = None) -> Iterator[Dict[str, Any]]:
        """Lee archivo JSONL línea por línea."""
        try:
            with open(self.file_path, 'r', encoding=self.encoding) as f:
                for line_num, line in enumerate(f, 1):
                    if max_rows and line_num > max_rows:
                        break
                    
                    line = line.strip()
                    if line:
                        try:
                            yield json.loads(line)
                        except json.JSONDecodeError as e:
                            logger.warning(f"Error en línea {line_num}: {e}")
                            continue
                            
        except Exception as e:
            raise ProcessingError(f"Error leyendo JSONL: {e}")
    
    def get_headers(self) -> List[str]:
        """Obtiene cabeceras del JSONL."""
        try:
            # Leer una muestra para obtener claves
            sample_data = list(self.read(max_rows=10))
            
            if not sample_data:
                return []
            
            # Obtener todas las claves únicas
            headers = set()
            for item in sample_data:
                if isinstance(item, dict):
                    headers.update(item.keys())
            
            return sorted(list(headers))
            
        except Exception as e:
            raise ProcessingError(f"Error obteniendo cabeceras JSONL: {e}")
    
    def count_rows(self) -> int:
        """Cuenta líneas del JSONL."""
        try:
            with open(self.file_path, 'r', encoding=self.encoding) as f:
                return sum(1 for line in f if line.strip())
        except Exception as e:
            raise ProcessingError(f"Error contando filas JSONL: {e}")

def create_file_reader(file_path: str, file_type: FileType = None, **kwargs) -> BaseFileReader:
    """
    Factory para crear lectores de archivos.
    
    Args:
        file_path: Ruta del archivo
        file_type: Tipo de archivo (se detecta si no se proporciona)
        **kwargs: Argumentos adicionales para el lector
        
    Returns:
        Instancia del lector apropiado
    """
    if file_type is None:
        file_type = detect_file_type(file_path)
    
    readers = {
        FileType.CSV: CSVReader,
        FileType.TSV: lambda path, **kw: CSVReader(path, delimiter='\t', **kw),
        FileType.EXCEL: ExcelReader,
        FileType.JSON: JSONReader,
        FileType.JSONL: JSONLReader
    }
    
    if file_type not in readers:
        raise ImportError(f"Tipo de archivo no soportado: {file_type}")
    
    return readers[file_type](file_path, **kwargs)

# ==============================================================================
# VALIDADORES DE DATOS
# ==============================================================================

class DataValidator:
    """Validador de datos para importación."""
    
    def __init__(self, validation_level: ValidationLevel = ValidationLevel.BASIC):
        self.validation_level = validation_level
        self.business_sectors = BUSINESS_SECTORS
        self.date_formats = IMPORT_CONFIG['date_formats']
        self.boolean_values = IMPORT_CONFIG['boolean_values']
        self.null_values = IMPORT_CONFIG['null_values']
    
    def validate_row(self, row_data: Dict[str, Any], 
                    field_mappings: List[FieldMapping],
                    row_number: int) -> Tuple[Dict[str, Any], List[str]]:
        """
        Valida una fila de datos.
        
        Args:
            row_data: Datos de la fila
            field_mappings: Mapeos de campos
            row_number: Número de fila
            
        Returns:
            Tupla (datos_validados, lista_errores)
        """
        validated_data = {}
        errors = []
        
        for mapping in field_mappings:
            source_field = mapping.source_field
            target_field = mapping.target_field
            
            # Obtener valor
            raw_value = row_data.get(source_field)
            
            try:
                # Validar si es requerido
                if mapping.required and self._is_null_value(raw_value):
                    errors.append(f"Campo requerido '{source_field}' está vacío")
                    continue
                
                # Aplicar valor por defecto si es necesario
                if self._is_null_value(raw_value) and mapping.default_value is not None:
                    raw_value = mapping.default_value
                
                # Transformar valor
                if mapping.transformer:
                    try:
                        processed_value = mapping.transformer(raw_value)
                    except Exception as e:
                        errors.append(f"Error en transformación de '{source_field}': {e}")
                        continue
                else:
                    processed_value = self._convert_data_type(raw_value, mapping.data_type)
                
                # Validar con validador personalizado
                if mapping.validator:
                    try:
                        is_valid, error_msg = mapping.validator(processed_value)
                        if not is_valid:
                            errors.append(f"Validación falló en '{source_field}': {error_msg}")
                            continue
                    except Exception as e:
                        errors.append(f"Error en validador de '{source_field}': {e}")
                        continue
                
                # Validaciones específicas
                validation_error = self._validate_by_field_name(target_field, processed_value)
                if validation_error:
                    errors.append(f"Error en '{source_field}': {validation_error}")
                    continue
                
                validated_data[target_field] = processed_value
                
            except Exception as e:
                errors.append(f"Error procesando '{source_field}': {e}")
        
        return validated_data, errors
    
    def _is_null_value(self, value: Any) -> bool:
        """Verifica si un valor se considera nulo."""
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip() in self.null_values
        return False
    
    def _convert_data_type(self, value: Any, data_type: str) -> Any:
        """Convierte valor al tipo de dato especificado."""
        if self._is_null_value(value):
            return None
        
        if isinstance(value, str):
            value = value.strip()
        
        try:
            if data_type == 'string':
                return str(value) if value is not None else None
            
            elif data_type == 'integer':
                if isinstance(value, str) and value == '':
                    return None
                return int(float(value))  # Manejar "123.0"
            
            elif data_type == 'float':
                if isinstance(value, str) and value == '':
                    return None
                return float(value)
            
            elif data_type == 'decimal':
                if isinstance(value, str) and value == '':
                    return None
                return Decimal(str(value))
            
            elif data_type == 'boolean':
                return self._convert_to_boolean(value)
            
            elif data_type == 'date':
                return self._convert_to_date(value)
            
            elif data_type == 'datetime':
                return self._convert_to_datetime(value)
            
            elif data_type == 'email':
                email = str(value).lower().strip()
                if '@' not in email:
                    raise ValueError("Formato de email inválido")
                return email
            
            elif data_type == 'url':
                url = str(value).strip()
                if not (url.startswith('http://') or url.startswith('https://')):
                    url = 'https://' + url
                return url
            
            elif data_type == 'phone':
                return self._clean_phone_number(str(value))
            
            else:
                return str(value)
                
        except Exception as e:
            raise ValueError(f"No se pudo convertir '{value}' a {data_type}: {e}")
    
    def _convert_to_boolean(self, value: Any) -> bool:
        """Convierte valor a booleano."""
        if isinstance(value, bool):
            return value
        
        if isinstance(value, (int, float)):
            return bool(value)
        
        if isinstance(value, str):
            value_lower = value.lower().strip()
            
            # Valores verdaderos
            for true_val in self.boolean_values['true']:
                if isinstance(true_val, str) and value_lower == true_val.lower():
                    return True
            
            # Valores falsos
            for false_val in self.boolean_values['false']:
                if isinstance(false_val, str) and value_lower == false_val.lower():
                    return False
        
        # Si no coincide, intentar conversión numérica
        try:
            return bool(float(value))
        except:
            raise ValueError(f"No se pudo convertir '{value}' a booleano")
    
    def _convert_to_date(self, value: Any) -> date:
        """Convierte valor a fecha."""
        if isinstance(value, date):
            return value
        
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, str):
            value = value.strip()
            
            for date_format in self.date_formats:
                try:
                    return datetime.strptime(value, date_format).date()
                except ValueError:
                    continue
        
        raise ValueError(f"No se pudo convertir '{value}' a fecha")
    
    def _convert_to_datetime(self, value: Any) -> datetime:
        """Convierte valor a datetime."""
        if isinstance(value, datetime):
            return value
        
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        
        if isinstance(value, str):
            value = value.strip()
            
            for date_format in self.date_formats:
                try:
                    return datetime.strptime(value, date_format)
                except ValueError:
                    continue
        
        raise ValueError(f"No se pudo convertir '{value}' a datetime")
    
    def _clean_phone_number(self, phone: str) -> str:
        """Limpia número de teléfono."""
        # Remover caracteres no numéricos excepto +
        cleaned = re.sub(r'[^\d+]', '', phone)
        
        # Validaciones básicas para Colombia
        if cleaned.startswith('+57'):
            if len(cleaned) == 13:  # +57 + 10 dígitos
                return cleaned
        elif cleaned.startswith('57'):
            if len(cleaned) == 12:  # 57 + 10 dígitos
                return '+' + cleaned
        elif len(cleaned) == 10:  # 10 dígitos
            return '+57' + cleaned
        
        return cleaned
    
    def _validate_by_field_name(self, field_name: str, value: Any) -> Optional[str]:
        """Validaciones específicas por nombre de campo."""
        if value is None:
            return None
        
        field_lower = field_name.lower()
        
        # Validar sector empresarial
        if 'sector' in field_lower and isinstance(value, str):
            if value.lower() not in self.business_sectors:
                return f"Sector '{value}' no es válido"
        
        # Validar email
        if 'email' in field_lower and isinstance(value, str):
            if '@' not in value or '.' not in value.split('@')[1]:
                return f"Email '{value}' tiene formato inválido"
        
        # Validar URLs
        if 'url' in field_lower or 'website' in field_lower:
            if isinstance(value, str) and value:
                if not (value.startswith('http://') or value.startswith('https://')):
                    return f"URL '{value}' debe incluir protocolo (http/https)"
        
        # Validar presupuesto/inversión
        if any(keyword in field_lower for keyword in ['presupuesto', 'budget', 'inversion', 'investment']):
            if isinstance(value, (int, float, Decimal)):
                if value < 0:
                    return f"El valor monetario no puede ser negativo: {value}"
        
        # Validar porcentajes
        if 'percentage' in field_lower or 'progreso' in field_lower:
            if isinstance(value, (int, float)):
                if not 0 <= value <= 100:
                    return f"Porcentaje debe estar entre 0 y 100: {value}"
        
        return None

# ==============================================================================
# TRANSFORMADORES DE DATOS
# ==============================================================================

class DataTransformer:
    """Transformador de datos para normalización."""
    
    @staticmethod
    def normalize_business_name(name: str) -> str:
        """Normaliza nombre de empresa."""
        if not isinstance(name, str):
            return str(name) if name else ''
        
        # Limpiar y normalizar
        normalized = name.strip()
        
        # Remover sufijos empresariales comunes
        suffixes = ['S.A.S.', 'SAS', 'S.A.', 'SA', 'LTDA', 'LTDA.', 'E.U.', 'EU']
        for suffix in suffixes:
            if normalized.upper().endswith(suffix):
                normalized = normalized[:-len(suffix)].strip()
                break
        
        return normalized
    
    @staticmethod
    def normalize_sector(sector: str) -> str:
        """Normaliza sector empresarial."""
        if not isinstance(sector, str):
            return str(sector) if sector else ''
        
        sector_normalized = sector.lower().strip()
        
        # Mapeo de sectores comunes
        sector_mapping = {
            'tech': 'tecnologia',
            'technology': 'tecnologia',
            'it': 'tecnologia',
            'software': 'tecnologia',
            'agro': 'agricultura',
            'farm': 'agricultura',
            'manufacturing': 'manufactura',
            'industry': 'manufactura',
            'retail': 'comercio',
            'commerce': 'comercio',
            'sales': 'comercio',
            'tourism': 'turismo',
            'travel': 'turismo',
            'education': 'educacion',
            'teaching': 'educacion',
            'health': 'salud',
            'healthcare': 'salud',
            'medical': 'salud',
            'construction': 'construccion',
            'building': 'construccion',
            'transport': 'transporte',
            'logistics': 'transporte',
            'energy': 'energia',
            'power': 'energia',
            'finance': 'finanzas',
            'banking': 'finanzas',
            'food': 'alimentario',
            'restaurant': 'alimentario',
            'textile': 'textil',
            'fashion': 'textil',
            'creative': 'creativos',
            'design': 'creativos',
            'mining': 'mineria',
            'telecom': 'telecomunicaciones',
            'consulting': 'consultoria',
            'advisory': 'consultoria'
        }
        
        return sector_mapping.get(sector_normalized, sector_normalized)
    
    @staticmethod
    def normalize_country(country: str) -> str:
        """Normaliza nombre de país."""
        if not isinstance(country, str):
            return str(country) if country else ''
        
        country_mapping = {
            'colombia': 'Colombia',
            'co': 'Colombia',
            'col': 'Colombia',
            'venezuela': 'Venezuela',
            've': 'Venezuela',
            'ven': 'Venezuela',
            'ecuador': 'Ecuador',
            'ec': 'Ecuador',
            'ecu': 'Ecuador',
            'peru': 'Perú',
            'pe': 'Perú',
            'per': 'Perú',
            'brasil': 'Brasil',
            'brazil': 'Brasil',
            'br': 'Brasil',
            'bra': 'Brasil',
            'argentina': 'Argentina',
            'ar': 'Argentina',
            'arg': 'Argentina',
            'chile': 'Chile',
            'cl': 'Chile',
            'chi': 'Chile'
        }
        
        country_normalized = country.lower().strip()
        return country_mapping.get(country_normalized, country.title())
    
    @staticmethod
    def normalize_education_level(education: str) -> str:
        """Normaliza nivel educativo."""
        if not isinstance(education, str):
            return str(education) if education else ''
        
        education_mapping = {
            'bachiller': 'Bachillerato',
            'bachillerato': 'Bachillerato',
            'high school': 'Bachillerato',
            'tecnico': 'Técnico',
            'technical': 'Técnico',
            'tecnologo': 'Tecnólogo',
            'technologist': 'Tecnólogo',
            'universitario': 'Universitario',
            'university': 'Universitario',
            'pregrado': 'Universitario',
            'bachelor': 'Universitario',
            'licenciatura': 'Universitario',
            'especializacion': 'Especialización',
            'specialization': 'Especialización',
            'maestria': 'Maestría',
            'master': 'Maestría',
            'mba': 'MBA',
            'doctorado': 'Doctorado',
            'doctorate': 'Doctorado',
            'phd': 'Doctorado'
        }
        
        education_normalized = education.lower().strip()
        return education_mapping.get(education_normalized, education.title())

# ==============================================================================
# MANAGER PRINCIPAL DE IMPORTACIÓN
# ==============================================================================

class ImportManager:
    """Manager principal para operaciones de importación."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = IMPORT_CONFIG.copy()
        if config:
            self.config.update(config)
        
        self.validator = DataValidator()
        self.transformer = DataTransformer()
        
        logger.info("ImportManager inicializado")
    
    def preview_data(self, file_path: str, 
                    max_rows: int = None,
                    field_mapping: Dict[str, str] = None) -> Dict[str, Any]:
        """
        Previsualiza datos del archivo sin importar.
        
        Args:
            file_path: Ruta del archivo
            max_rows: Número máximo de filas a previsualizar
            field_mapping: Mapeo de campos opcional
            
        Returns:
            Diccionario con preview de datos
            
        Examples:
            >>> manager = ImportManager()
            >>> preview = manager.preview_data('entrepreneurs.csv', max_rows=5)
            >>> print(f"Cabeceras: {preview['headers']}")
            >>> print(f"Filas: {len(preview['sample_data'])}")
        """
        if max_rows is None:
            max_rows = self.config['max_preview_rows']
        
        try:
            # Validar archivo
            file_info = validate_file(file_path)
            
            # Crear lector
            reader = create_file_reader(file_path, file_info['type'])
            
            # Obtener cabeceras
            headers = reader.get_headers()
            
            # Leer datos de muestra
            sample_data = list(reader.read(max_rows=max_rows))
            
            # Aplicar mapeo si se proporciona
            if field_mapping:
                mapped_headers = []
                mapped_data = []
                
                for row in sample_data:
                    mapped_row = {}
                    for source_field, target_field in field_mapping.items():
                        if source_field in row:
                            mapped_row[target_field] = row[source_field]
                    mapped_data.append(mapped_row)
                
                mapped_headers = list(field_mapping.values())
            else:
                mapped_headers = headers
                mapped_data = sample_data
            
            # Análisis de datos
            analysis = self._analyze_data(sample_data, headers)
            
            return {
                'file_info': file_info,
                'headers': headers,
                'mapped_headers': mapped_headers,
                'sample_data': mapped_data,
                'total_rows': reader.count_rows(),
                'analysis': analysis
            }
            
        except Exception as e:
            logger.error(f"Error en preview: {e}")
            raise ImportError(f"Error previsualizando archivo: {e}")
    
    def import_data(self, config: ImportConfig) -> ImportResult:
        """
        Importa datos según la configuración especificada.
        
        Args:
            config: Configuración de importación
            
        Returns:
            Resultado de la importación
            
        Examples:
            >>> config = ImportConfig(
            ...     file_path='entrepreneurs.csv',
            ...     target_model='Entrepreneur',
            ...     field_mappings=[
            ...         FieldMapping('nombre', 'name', 'string', required=True),
            ...         FieldMapping('email', 'email', 'email', required=True)
            ...     ]
            ... )
            >>> result = manager.import_data(config)
            >>> print(f"Éxito: {result.get_success_rate():.1f}%")
        """
        start_time = time.time()
        
        # Inicializar resultado
        result = ImportResult(
            status=ImportStatus.PROCESSING,
            field_mapping={fm.source_field: fm.target_field for fm in config.field_mappings}
        )
        
        try:
            # Validar archivo
            file_info = validate_file(config.file_path)
            result.file_info = file_info
            
            # Crear lector
            reader = create_file_reader(config.file_path, file_info['type'])
            
            # Contar filas totales
            result.total_rows = reader.count_rows()
            
            if config.max_rows:
                result.total_rows = min(result.total_rows, config.max_rows)
            
            logger.info(f"Iniciando importación de {result.total_rows} filas")
            
            # Procesar datos
            if config.preview_only:
                result = self._process_preview(reader, config, result)
            else:
                result = self._process_import(reader, config, result)
            
            result.processing_time = time.time() - start_time
            result.status = ImportStatus.COMPLETED
            
            logger.info(f"Importación completada en {result.processing_time:.2f}s")
            logger.info(f"Éxito: {result.get_success_rate():.1f}% ({result.successful_rows}/{result.total_rows})")
            
        except Exception as e:
            result.status = ImportStatus.FAILED
            result.processing_time = time.time() - start_time
            logger.error(f"Error en importación: {e}")
            result.add_error(0, f"Error fatal: {e}")
        
        return result
    
    def _process_preview(self, reader: BaseFileReader, 
                        config: ImportConfig, 
                        result: ImportResult) -> ImportResult:
        """Procesa preview de datos."""
        max_rows = min(config.max_rows or 100, self.config['max_preview_rows'])
        
        for row_num, row_data in enumerate(reader.read(max_rows=max_rows), 1):
            try:
                # Validar y transformar
                validated_data, errors = self.validator.validate_row(
                    row_data, config.field_mappings, row_num
                )
                
                if errors:
                    for error in errors:
                        result.add_error(row_num, error)
                else:
                    result.successful_rows += 1
                    result.imported_data.append(validated_data)
                
                result.processed_rows += 1
                
                # Callback de progreso
                if config.progress_callback and row_num % self.config['progress_callback_interval'] == 0:
                    config.progress_callback(row_num, result.total_rows)
                
            except Exception as e:
                result.add_error(row_num, f"Error procesando fila: {e}")
                result.processed_rows += 1
        
        return result
    
    def _process_import(self, reader: BaseFileReader, 
                      config: ImportConfig, 
                      result: ImportResult) -> ImportResult:
        """Procesa importación completa de datos."""
        batch = []
        batch_num = 0
        
        for row_num, row_data in enumerate(reader.read(max_rows=config.max_rows), 1):
            try:
                # Validar y transformar
                validated_data, errors = self.validator.validate_row(
                    row_data, config.field_mappings, row_num
                )
                
                if errors:
                    for error in errors:
                        result.add_error(row_num, error)
                    
                    # Si hay rollback activado y muchos errores, abortar
                    if config.rollback_on_error and len(result.errors) > result.total_rows * 0.1:
                        raise ProcessingError("Demasiados errores, abortando importación")
                else:
                    batch.append(validated_data)
                    result.successful_rows += 1
                
                result.processed_rows += 1
                
                # Procesar lote cuando esté lleno
                if len(batch) >= config.batch_size:
                    batch_num += 1
                    self._process_batch(batch, config, result, batch_num)
                    batch = []
                
                # Callback de progreso
                if config.progress_callback and row_num % self.config['progress_callback_interval'] == 0:
                    config.progress_callback(row_num, result.total_rows)
                
            except Exception as e:
                result.add_error(row_num, f"Error procesando fila: {e}")
                result.processed_rows += 1
        
        # Procesar último lote
        if batch:
            batch_num += 1
            self._process_batch(batch, config, result, batch_num)
        
        return result
    
    def _process_batch(self, batch: List[Dict[str, Any]], 
                      config: ImportConfig, 
                      result: ImportResult, 
                      batch_num: int):
        """Procesa un lote de datos."""
        try:
            logger.debug(f"Procesando lote {batch_num} con {len(batch)} registros")
            
            # Aquí se haría la inserción en base de datos
            # if config.target_model and SQLALCHEMY_AVAILABLE:
            #     self._insert_to_database(batch, config.target_model)
            
            # Por ahora, solo añadimos a los datos importados
            result.imported_data.extend(batch)
            
        except Exception as e:
            logger.error(f"Error procesando lote {batch_num}: {e}")
            # Marcar todos los registros del lote como fallidos
            for _ in batch:
                result.add_error(batch_num, f"Error en lote: {e}")
    
    def _analyze_data(self, sample_data: List[Dict[str, Any]], 
                     headers: List[str]) -> Dict[str, Any]:
        """Analiza muestra de datos para estadísticas."""
        if not sample_data:
            return {}
        
        analysis = {
            'row_count': len(sample_data),
            'column_count': len(headers),
            'columns': {}
        }
        
        for header in headers:
            values = [row.get(header) for row in sample_data]
            non_null_values = [v for v in values if v is not None and str(v).strip()]
            
            column_analysis = {
                'non_null_count': len(non_null_values),
                'null_count': len(values) - len(non_null_values),
                'null_percentage': ((len(values) - len(non_null_values)) / len(values)) * 100,
                'sample_values': non_null_values[:5],  # Primeros 5 valores
                'suggested_type': self._suggest_data_type(non_null_values)
            }
            
            analysis['columns'][header] = column_analysis
        
        return analysis
    
    def _suggest_data_type(self, values: List[Any]) -> str:
        """Sugiere tipo de dato basado en valores."""
        if not values:
            return 'string'
        
        # Contar tipos exitosos
        type_counts = {
            'integer': 0,
            'float': 0,
            'boolean': 0,
            'date': 0,
            'email': 0,
            'string': 0
        }
        
        for value in values[:20]:  # Analizar primeros 20 valores
            if value is None:
                continue
            
            str_value = str(value).strip()
            
            # Probar integer
            try:
                int(float(str_value))
                type_counts['integer'] += 1
                continue
            except:
                pass
            
            # Probar float
            try:
                float(str_value)
                type_counts['float'] += 1
                continue
            except:
                pass
            
            # Probar boolean
            if str_value.lower() in ['true', 'false', 'yes', 'no', '1', '0', 'sí', 'no']:
                type_counts['boolean'] += 1
                continue
            
            # Probar fecha
            for date_format in self.config['date_formats'][:5]:  # Probar primeros 5 formatos
                try:
                    datetime.strptime(str_value, date_format)
                    type_counts['date'] += 1
                    break
                except:
                    continue
            else:
                # Probar email
                if '@' in str_value and '.' in str_value:
                    type_counts['email'] += 1
                else:
                    type_counts['string'] += 1
        
        # Retornar tipo más común
        return max(type_counts, key=type_counts.get)
    
    def get_field_mapping_suggestions(self, file_path: str, 
                                    target_model: str = None) -> Dict[str, str]:
        """
        Sugiere mapeo de campos basado en nombres comunes.
        
        Args:
            file_path: Ruta del archivo
            target_model: Modelo objetivo (opcional)
            
        Returns:
            Diccionario con sugerencias de mapeo
        """
        try:
            # Obtener cabeceras del archivo
            reader = create_file_reader(file_path)
            headers = reader.get_headers()
            
            # Normalizar cabeceras
            normalized_headers = [header.lower().strip() for header in headers]
            
            # Seleccionar mapeo base según el modelo
            if target_model and target_model.lower() == 'entrepreneur':
                base_mapping = ENTREPRENEUR_FIELD_MAPPING
            elif target_model and target_model.lower() == 'project':
                base_mapping = PROJECT_FIELD_MAPPING
            else:
                base_mapping = {**ENTREPRENEUR_FIELD_MAPPING, **PROJECT_FIELD_MAPPING}
            
            # Generar sugerencias
            suggestions = {}
            for header, normalized in zip(headers, normalized_headers):
                for source_pattern, target_field in base_mapping.items():
                    if source_pattern in normalized or normalized in source_pattern:
                        suggestions[header] = target_field
                        break
            
            return suggestions
            
        except Exception as e:
            logger.error(f"Error generando sugerencias de mapeo: {e}")
            return {}

# ==============================================================================
# FUNCIONES DE CONVENIENCIA
# ==============================================================================

def import_from_csv(file_path: str, 
                   target_model: str = None,
                   field_mapping: Dict[str, str] = None,
                   **kwargs) -> ImportResult:
    """
    Función de conveniencia para importar desde CSV.
    
    Args:
        file_path: Ruta del archivo CSV
        target_model: Modelo objetivo
        field_mapping: Mapeo de campos
        **kwargs: Argumentos adicionales
        
    Returns:
        Resultado de importación
        
    Examples:
        >>> result = import_from_csv(
        ...     'entrepreneurs.csv',
        ...     target_model='Entrepreneur',
        ...     field_mapping={'nombre': 'name', 'email': 'email'}
        ... )
        >>> print(f"Importados: {result.successful_rows} registros")
    """
    manager = ImportManager()
    
    # Crear mapeos de campos
    field_mappings = []
    if field_mapping:
        for source, target in field_mapping.items():
            field_mappings.append(FieldMapping(
                source_field=source,
                target_field=target,
                required=target in ['name', 'email']  # Campos básicos requeridos
            ))
    
    # Configurar importación
    config = ImportConfig(
        file_path=file_path,
        target_model=target_model,
        field_mappings=field_mappings,
        **kwargs
    )
    
    return manager.import_data(config)

def import_from_excel(file_path: str, 
                     sheet_name: str = None,
                     target_model: str = None,
                     field_mapping: Dict[str, str] = None,
                     **kwargs) -> ImportResult:
    """
    Función de conveniencia para importar desde Excel.
    
    Args:
        file_path: Ruta del archivo Excel
        sheet_name: Nombre de la hoja
        target_model: Modelo objetivo
        field_mapping: Mapeo de campos
        **kwargs: Argumentos adicionales
        
    Returns:
        Resultado de importación
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está disponible para archivos Excel")
    
    manager = ImportManager()
    
    # Crear mapeos de campos
    field_mappings = []
    if field_mapping:
        for source, target in field_mapping.items():
            field_mappings.append(FieldMapping(
                source_field=source,
                target_field=target
            ))
    
    # Configurar importación
    config = ImportConfig(
        file_path=file_path,
        target_model=target_model,
        field_mappings=field_mappings,
        **kwargs
    )
    
    return manager.import_data(config)

def import_from_json(file_path: str, 
                    target_model: str = None,
                    field_mapping: Dict[str, str] = None,
                    **kwargs) -> ImportResult:
    """
    Función de conveniencia para importar desde JSON.
    
    Args:
        file_path: Ruta del archivo JSON
        target_model: Modelo objetivo
        field_mapping: Mapeo de campos
        **kwargs: Argumentos adicionales
        
    Returns:
        Resultado de importación
    """
    manager = ImportManager()
    
    # Crear mapeos de campos
    field_mappings = []
    if field_mapping:
        for source, target in field_mapping.items():
            field_mappings.append(FieldMapping(
                source_field=source,
                target_field=target
            ))
    
    # Configurar importación
    config = ImportConfig(
        file_path=file_path,
        target_model=target_model,
        field_mappings=field_mappings,
        **kwargs
    )
    
    return manager.import_data(config)

def validate_import_data(file_path: str, 
                        field_mapping: Dict[str, str] = None,
                        max_rows: int = 100) -> Dict[str, Any]:
    """
    Valida datos de importación sin importar.
    
    Args:
        file_path: Ruta del archivo
        field_mapping: Mapeo de campos
        max_rows: Número máximo de filas a validar
        
    Returns:
        Resultado de validación
        
    Examples:
        >>> validation = validate_import_data('data.csv', {'nombre': 'name'})
        >>> if validation['has_errors']:
        ...     print(f"Errores encontrados: {len(validation['errors'])}")
    """
    manager = ImportManager()
    
    try:
        # Preview con validación
        preview = manager.preview_data(file_path, max_rows, field_mapping)
        
        # Crear configuración temporal para validación
        field_mappings = []
        if field_mapping:
            for source, target in field_mapping.items():
                field_mappings.append(FieldMapping(source, target))
        
        config = ImportConfig(
            file_path=file_path,
            field_mappings=field_mappings,
            max_rows=max_rows,
            preview_only=True
        )
        
        # Ejecutar validación
        result = manager.import_data(config)
        
        return {
            'has_errors': len(result.errors) > 0,
            'errors': result.errors,
            'warnings': result.warnings,
            'success_rate': result.get_success_rate(),
            'total_rows': result.total_rows,
            'preview': preview
        }
        
    except Exception as e:
        return {
            'has_errors': True,
            'errors': [{'row': 0, 'message': str(e)}],
            'warnings': [],
            'success_rate': 0.0,
            'total_rows': 0,
            'preview': None
        }

# ==============================================================================
# INSTANCIAS GLOBALES Y CONFIGURACIÓN
# ==============================================================================

# Instancia global del import manager
import_manager = ImportManager()

def get_import_config() -> Dict[str, Any]:
    """Obtiene configuración actual de importación."""
    return import_manager.config.copy()

def configure_import(**kwargs):
    """Configura utilidades de importación globalmente."""
    import_manager.config.update(kwargs)

def get_supported_formats() -> List[str]:
    """Obtiene formatos de archivo soportados."""
    formats = ['csv', 'tsv', 'json', 'jsonl']
    
    if OPENPYXL_AVAILABLE:
        formats.extend(['xlsx', 'xls'])
    
    if XML_AVAILABLE:
        formats.append('xml')
    
    return formats

def get_import_statistics() -> Dict[str, Any]:
    """Obtiene estadísticas del sistema de importación."""
    return {
        'supported_formats': get_supported_formats(),
        'dependencies': {
            'openpyxl': OPENPYXL_AVAILABLE,
            'xlrd': XLRD_AVAILABLE,
            'xml': XML_AVAILABLE,
            'sqlalchemy': SQLALCHEMY_AVAILABLE
        },
        'config': get_import_config(),
        'business_sectors': sorted(list(BUSINESS_SECTORS)),
        'field_mappings': {
            'entrepreneur': ENTREPRENEUR_FIELD_MAPPING,
            'project': PROJECT_FIELD_MAPPING
        }
    }

# Logging de inicialización
available_formats = get_supported_formats()
logger.info(f"Módulo de importación inicializado - Formatos: {', '.join(available_formats)}")

if not OPENPYXL_AVAILABLE:
    logger.warning("openpyxl no disponible - archivos Excel no soportados")